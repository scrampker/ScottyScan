#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OpenVAS_25 (latest scan) -> Excel:
- Import most recent 'detailedresults.csv' under a root folder.
- Details tab is High/Medium only and date-stamped as OpenVAS_YYMMDD.
- Summary tab (OpenVAS_Summary) is always first and accumulates per-host totals over time:
    * high_count_now, medium_count_now, totals_now (H+M only for the current scan)
    * One per-scan H+M total column named:
        - `baseline-YYMMDD-totals` on the first run
        - `scan-YYMMDD-totals` on each subsequent scan
- Normalize IP (first IPv4) and Hostname (lowercase).
- Split CVEs to one row per CVE (explode).
- Remove duplicate findings (ip, hostname, port, nvt_oid, cve).
- Enrich from asset_tracker.xlsx (match by IP first; else hostname vs name/dns_name).
- Sheets written/updated each run:
    * OpenVAS_Summary (first sheet, accumulated)
    * OpenVAS_YYMMDD (HM-only details for this scan)
    * _meta (appended with run metadata)
Notes:
- Software_By_Host and Software_Summary are currently disabled.
"""

import argparse
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import sys

# -------------------------------
# OpenVAS CSV canonical header strings (to avoid magic strings)
# -------------------------------
OV_COL_IP            = "ip"
OV_COL_HOSTNAME      = "hostname"
OV_COL_PORT          = "port"
OV_COL_PORT_PROTOCOL = "port protocol"
OV_COL_CVSS          = "cvss"
OV_COL_SEVERITY      = "severity"
OV_COL_QOD           = "qod"
OV_COL_NVT_NAME      = "nvt name"
OV_COL_SUMMARY       = "summary"
OV_COL_SPECIFIC      = "specific result"
OV_COL_NVT_OID       = "nvt oid"
OV_COL_CVES          = "cves"
OV_COL_TIMESTAMP     = "timestamp"
OV_COL_AFFECTED_SW   = "affected software/os"
OV_COL_PRODUCT_DET   = "product detection result"

# -------------------------------
# Utils
# -------------------------------

IPV4_RE = re.compile(r"\b((?:\d{1,3}\.){3}\d{1,3})\b")
# Capture standard CPE URIs (allowing additional safe chars and dashes/underscores)
CPE_RE = re.compile(r"(cpe:/[aoh]:[^,\s\"')]+)", re.IGNORECASE)
# Common architecture tokens that sometimes appear as a mistaken CPE version
ARCH_TOKEN_RE = re.compile(r"^(x64|x86|amd64|arm64)$", re.IGNORECASE)

def first_ipv4(s):
    if s is None:
        return None
    m = IPV4_RE.search(str(s))
    return m.group(1) if m else None

def normalize_host(s):
    return str(s).strip().lower() if s is not None else ""

def newest_openvas_csv(root: Path) -> Path:
    """
    Return the most recent detailedresults.csv under root.
    If none found, raise FileNotFoundError with guidance and nearby hints
    (e.g., zipped scans or similarly named CSVs).
    """
    root = Path(root)
    candidates = list(root.rglob("detailedresults.csv"))
    if candidates:
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return candidates[0]

    # No direct hit – look for likely alternatives to help the user
    hint_csvs = []
    for pat in ("*detailed*results*.csv", "*openvas*.csv", "*results*.csv"):
        hint_csvs.extend(root.rglob(pat))
    hint_csvs = sorted({p for p in hint_csvs if p.is_file()})

    hint_zips = sorted({p for p in root.rglob("*.zip") if p.is_file()})

    msg = [f"No detailedresults.csv found under {root}"]
    if hint_csvs:
        msg.append("\nOther CSVs that look related:")
        for p in hint_csvs[:10]:
            msg.append(f"  - {p}")
        if len(hint_csvs) > 10:
            msg.append(f"  ... (+{len(hint_csvs)-10} more)")
    if hint_zips:
        msg.append("\nZip files detected — if your OpenVAS export is zipped, extract it or point to the folder that contains detailedresults.csv:")
        for p in hint_zips[:10]:
            msg.append(f"  - {p}")
        if len(hint_zips) > 10:
            msg.append(f"  ... (+{len(hint_zips)-10} more)")

    msg.append("\nTips:\n  • Use --openvas-root to point at the top-level scan folder.\n  • Ensure the file name is exactly 'detailedresults.csv'.\n  • If using a zip, unzip it so the CSV exists on disk.")
    raise FileNotFoundError("\n".join(msg))

def split_cves(val):
    if val is None or str(val).strip() == "":
        return [""]
    raw = str(val)
    parts = re.split(r"[,\s]+", raw.strip())
    parts = [p.strip() for p in parts if p.strip()]
    return parts or [""]

# -------------------------------
# Loaders
# -------------------------------

def load_asset_tracker(path: Path) -> pd.DataFrame:
    """
    Expect sheet 'asset_tracker' with at least:
      - name, dns_name, ip_address, status, location, notes
    Returns (df, ip_map, host_map)
    """
    df = pd.read_excel(path, sheet_name="asset_tracker", dtype=str)
    df = df.fillna("")
    df["name"] = df["name"].astype(str).str.strip().str.lower()
    df["dns_name"] = df["dns_name"].astype(str).str.strip().str.lower()
    df["ip_address"] = df["ip_address"].astype(str).str.strip().str.lower()

    ip_map = {}
    host_map = {}
    for _, r in df.iterrows():
        ip = r.get("ip_address", "")
        nm = r.get("name", "")
        dn = r.get("dns_name", "")
        if ip:
            ip_map[ip] = r
        if nm:
            host_map[nm] = r
        if dn:
            host_map[dn] = r
    return df, ip_map, host_map

def _resolve_cols(df, wanted_lowers):
    out = {}
    for target in wanted_lowers:
        hit = None
        for c in df.columns:
            if c.strip().lower() == target:
                hit = c
                break
        out[target] = hit
    return out

def load_openvas_csv(path: Path) -> pd.DataFrame:
    """
    Load OpenVAS detailedresults.csv.
    Return normalized table with canonical column names.
    """
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")

    cols = _resolve_cols(df, [
        OV_COL_IP, OV_COL_HOSTNAME, OV_COL_PORT, OV_COL_PORT_PROTOCOL, OV_COL_CVSS, OV_COL_SEVERITY, OV_COL_QOD,
        OV_COL_NVT_NAME, OV_COL_SUMMARY, OV_COL_SPECIFIC, OV_COL_NVT_OID, OV_COL_CVES, OV_COL_TIMESTAMP,
        OV_COL_AFFECTED_SW, OV_COL_PRODUCT_DET
    ])

    # Required
    for req in (OV_COL_IP, OV_COL_HOSTNAME, OV_COL_SEVERITY):
        if not cols[req]:
            raise RuntimeError(f"OpenVAS CSV missing required column: {req}")

    # Normalize IP + Hostname
    df["_ip_norm"] = df[cols[OV_COL_IP]].map(first_ipv4).fillna("")
    df["_host_norm"] = df[cols[OV_COL_HOSTNAME]].map(normalize_host)

    # Explode CVEs -> one row per CVE (use "" when none)
    df["_cve_list"] = df[cols[OV_COL_CVES]].map(split_cves) if cols[OV_COL_CVES] else [[""]] * len(df)
    df = df.explode("_cve_list", ignore_index=True)

    out = pd.DataFrame({
        "ip": df["_ip_norm"],
        "hostname": df["_host_norm"],
        "port": df[cols[OV_COL_PORT]] if cols[OV_COL_PORT] else "",
        "protocol": df[cols[OV_COL_PORT_PROTOCOL]] if cols[OV_COL_PORT_PROTOCOL] else "",
        "cvss": df[cols[OV_COL_CVSS]] if cols[OV_COL_CVSS] else "",
        "severity": df[cols[OV_COL_SEVERITY]],
        "qod": df[cols[OV_COL_QOD]] if cols[OV_COL_QOD] else "",
        "nvt_name": df[cols[OV_COL_NVT_NAME]] if cols[OV_COL_NVT_NAME] else "",
        "nvt_oid": df[cols[OV_COL_NVT_OID]] if cols[OV_COL_NVT_OID] else "",
        "cve": df["_cve_list"],
        "timestamp": df[cols[OV_COL_TIMESTAMP]] if cols[OV_COL_TIMESTAMP] else "",
        "affected_software_os": df[cols[OV_COL_AFFECTED_SW]] if cols[OV_COL_AFFECTED_SW] else "",
        "product_detection": df[cols[OV_COL_PRODUCT_DET]] if cols[OV_COL_PRODUCT_DET] else "",
        "summary": df[cols[OV_COL_SUMMARY]] if cols[OV_COL_SUMMARY] else "",
        "specific_result": df[cols[OV_COL_SPECIFIC]] if cols[OV_COL_SPECIFIC] else "",
    })

    # Coerce numeric fields so they are written as numbers in Excel
    out["cvss"] = pd.to_numeric(out["cvss"], errors="coerce")
    out["qod"] = pd.to_numeric(out["qod"], errors="coerce")

    # Deduplicate core vuln rows
    out = out.drop_duplicates(subset=["ip", "hostname", "port", "nvt_oid", "cve"]).reset_index(drop=True)
    return out

# -------------------------------
# Enrichment from asset_tracker
# -------------------------------

def enrich_with_assets(openvas_df: pd.DataFrame, ip_map, host_map) -> pd.DataFrame:
    """
    Add asset fields from asset_tracker by IP first, else hostname vs name/dns_name.
    """
    openvas_df = openvas_df.copy()
    openvas_df["asset_name"] = ""
    openvas_df["asset_dns_name"] = ""
    openvas_df["asset_status"] = ""
    openvas_df["asset_location"] = ""
    openvas_df["asset_source"] = ""
    openvas_df["asset_notes"] = ""

    for idx, row in openvas_df.iterrows():
        ip = str(row.get("ip") or "").strip().lower()
        host = str(row.get("hostname") or "").strip().lower()

        match = None
        if ip and ip in ip_map:
            match = ip_map[ip]
        elif host and host in host_map:
            match = host_map[host]

        if match is not None:
            openvas_df.at[idx, "asset_name"] = str(match.get("name", ""))
            openvas_df.at[idx, "asset_dns_name"] = str(match.get("dns_name", ""))
            openvas_df.at[idx, "asset_status"] = str(match.get("status", ""))
            openvas_df.at[idx, "asset_location"] = str(match.get("location", ""))
            openvas_df.at[idx, "asset_source"] = str(match.get("source", ""))
            openvas_df.at[idx, "asset_notes"] = str(match.get("notes", ""))

    return openvas_df

# --- Helper: lookup asset row for display host ---
def lookup_asset_for_display(host_key: str, ip_map, host_map):
    """Return an asset row (Series/dict) by matching the display host string
    against ip_address, name, or dns_name from the tracker (all lowercased).
    host_key is already a display token used in the summary (name|dns|ip).
    """
    if not host_key:
        return None
    k = str(host_key).strip().lower()
    if k in ip_map:
        return ip_map[k]
    if k in host_map:
        return host_map[k]
    return None

# -------------------------------
# Summary
# -------------------------------

def build_summary(openvas_df: pd.DataFrame) -> pd.DataFrame:
    """
    Roll up counts per host across all severities present (High/Medium/Low/Log/Info).
    """
    view = openvas_df.copy()

    def pick_display(r):
        nm = str(r.get("asset_name") or "").strip()
        if nm:
            return nm
        hn = str(r.get("hostname") or "").strip()
        if hn:
            return hn
        ip = str(r.get("ip") or "").strip()
        return ip if ip else "unknown"

    view["display_host"] = view.apply(pick_display, axis=1)
    view["severity_norm"] = view["severity"].astype(str).str.strip().str.lower()

    grp = view.groupby(["display_host", "severity_norm"]).size().unstack(fill_value=0)
    # Ensure all columns exist if missing
    for col in ("high", "medium", "low", "log", "info"):
        if col not in grp.columns:
            grp[col] = 0
    grp["total_findings"] = grp.sum(axis=1)

    grp = grp.reset_index().rename(columns={"display_host": "host"})

    firsts = view.groupby("display_host").agg({
        "asset_status": "first",
        "asset_location": "first",
        "asset_source": "first"
    }).reset_index().rename(columns={"display_host": "host"})

    summary = grp.merge(firsts, on="host", how="left")
    summary = summary.sort_values(
        by=["high", "medium", "low", "total_findings", "host"],
        ascending=[False, False, False, False, True]
    ).reset_index(drop=True)
    return summary

# -------------------------------
# Software Inventory / Summary
# -------------------------------

def extract_cpes_from_fields(row) -> list[tuple[str, str]]:
    """
    Return list of (cpe, source_field) tuples extracted from relevant text fields.
    Priority: affected_software_os -> product_detection -> specific_result -> summary.
    If no CPEs are found in a field, apply heuristic software parsing to emit pseudo-CPEs
    of the form 'cpe:/a:<vendor>:<product>:<version>' (part 'a' assumed for applications).
    """
    results: list[tuple[str, str]] = []

    def _scan_text(text: str, field: str):
        if not text:
            return []
        hits = []
        for m in CPE_RE.finditer(text):
            hits.append((m.group(1).lower(), field))
        if hits:
            return hits
        # Fallback: heuristic parse if no CPEs were found in this field
        for vendor, product, version in _fallback_parse_software(text):
            pseudo = f"cpe:/a:{vendor}:{product}:{version}".rstrip(":")
            hits.append((pseudo, field + "_heuristic"))
        return hits

    # Priority order
    for fld in ("affected_software_os", "product_detection", "specific_result", "summary"):
        val = row.get(fld)
        results.extend(_scan_text(val, fld))

    return results

def parse_cpe(cpe: str):
    """
    Parse a CPE like cpe:/a:vendor:product:version:...
    Return (part, vendor, product, version) with best-effort extraction.
    Normalize tokens to lowercase; strip blanks. Remove arch tokens from version.
    """
    if not cpe or not cpe.startswith("cpe:/"):
        return ("", "", "", "")
    body = cpe[5:]
    parts = body.split(":")
    part = (parts[0] if len(parts) > 0 else "").strip().lower()
    vendor = (parts[1] if len(parts) > 1 else "").strip().lower()
    product = (parts[2] if len(parts) > 2 else "").strip().lower()
    version = (parts[3] if len(parts) > 3 else "").strip()
    # Treat architecture tokens as not-a-version
    if ARCH_TOKEN_RE.match(version or ""):
        version = ""
    return (part, vendor, product, version)

# Heuristic fallback parser for software info (pseudo-CPEs)
def _fallback_parse_software(text: str):
    """Best-effort extraction of (vendor, product, version) when no CPE is present.
    Looks for patterns like 'Vendor Product X.Y[.Z]' in Affected Software/OS, Summary, or Specific Result.
    Returns a list of tuples (vendor, product, version).
    """
    if not text:
        return []
    s = str(text)
    # Common patterns we see in OpenVAS content
    patterns = [
        # 7-Zip 9.20
        re.compile(r"\b(7[- ]?zip)\s+(\d[\w\.-]*)", re.IGNORECASE),
        # Apache Tomcat 8.5.71
        re.compile(r"\b(apache)\s+(tomcat)\s+(\d[\w\.-]*)", re.IGNORECASE),
        # Microsoft ASP.NET Core 5.0 / 3.1.32 / 8.0.3
        re.compile(r"\b(microsoft)\s+(asp\.net(?:\s+core)?)\s+(\d[\w\.-]*)", re.IGNORECASE),
        # Adobe Acrobat Reader 11.0.10 / Adobe Flash Player
        re.compile(r"\b(adobe)\s+(acrobat(?:\s+reader)?|flash(?:\s+player)?)\s*(\d[\w\.-]*)?", re.IGNORECASE),
    ]
    results = []
    for pat in patterns:
        for m in pat.finditer(s):
            vendor = (m.group(1) or "").strip().lower()
            product = (m.group(2) or "").strip().lower().replace(" ", "_")
            version = (m.group(3) or "").strip()
            if ARCH_TOKEN_RE.match(version or ""):
                version = ""
            results.append((vendor, product, version))
    return results

def build_software_tables(openvas_df_for_sw: pd.DataFrame, openvas_df_for_keys: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build Software_By_Host (detailed) and Software_Summary (rollup).
    openvas_df_for_sw: the dataset we scan for CPEs (filtered or full, depending on --software-all).
    openvas_df_for_keys: used to compute display_host / fallback keys (usually the High/Medium enriched DF).
    """
    # Map for display host using enriched keys (prefer asset_name, else hostname, else ip)
    key_rows = []
    for _, r in openvas_df_for_keys.iterrows():
        nm = str(r.get("asset_name") or "").strip()
        hn = str(r.get("hostname") or "").strip()
        ip = str(r.get("ip") or "").strip()
        disp = nm or hn or ip or "unknown"
        key_rows.append((ip.lower(), hn.lower(), disp))
    # last-one-wins maps
    ip_disp = {}
    host_disp = {}
    for ip, hn, disp in key_rows:
        if ip:
            ip_disp[ip] = disp
        if hn:
            host_disp[hn] = disp

    rows = []
    for _, r in openvas_df_for_sw.iterrows():
        ip = str(r.get("ip") or "").strip().lower()
        hn = str(r.get("hostname") or "").strip().lower()
        disp = ip_disp.get(ip) or host_disp.get(hn) or hn or ip or "unknown"

        for cpe, src in extract_cpes_from_fields(r):
            part, vendor, product, version = parse_cpe(cpe)
            rows.append({
                "host": disp,
                "ip": ip,
                "hostname": hn,
                "cpe": cpe,
                "part": part,
                "vendor": vendor,
                "product": product,
                "version": version,
                "source_field": src
            })

    if not rows:
        sw_by_host = pd.DataFrame(columns=[
            "host","ip","hostname","cpe","part","vendor","product","version","source_field"
        ])
        sw_summary = pd.DataFrame(columns=[
            "software_key","distinct_hosts","occurrences"
        ])
        return sw_by_host, sw_summary

    sw_by_host = pd.DataFrame(rows)
    # Dedup per host/cpe to avoid double counting the same hit
    sw_by_host = sw_by_host.drop_duplicates(subset=["host", "cpe"]).reset_index(drop=True)

    # Software key: vendor:product (blank-safe), normalize double-colons
    sw_by_host["software_key"] = (
        sw_by_host["vendor"].fillna("") + ":" + sw_by_host["product"].fillna("")
    ).str.replace(r"^:\|:$", "", regex=True).str.strip(":")

    # Summary: distinct hosts using each software (vendor:product only)
    agg = sw_by_host.groupby(["software_key"]).agg(
        distinct_hosts=("host", "nunique"),
        occurrences=("cpe", "count")
    ).reset_index().sort_values(
        by=["distinct_hosts", "occurrences", "software_key"],
        ascending=[False, False, True]
    )

    return sw_by_host, agg

# -------------------------------
# Driver
# -------------------------------

def build_openvas_book(asset_tracker_xlsx: Path, openvas_root: Path, output_xlsx: Path):
    # Load assets and latest CSV
    _, ip_map, host_map = load_asset_tracker(asset_tracker_xlsx)
    latest = newest_openvas_csv(openvas_root)

    # Load the full OpenVAS CSV (no severity filtering)
    ov_full = load_openvas_csv(latest)
    # Build two datasets: full (for summary) and HM-only (for details)
    sev_norm = ov_full["severity"].astype(str).str.strip().str.lower()
    ov_hm = ov_full.loc[sev_norm.isin({"high", "medium"})].copy()

    # Enrich both datasets with assets
    ov_full = enrich_with_assets(ov_full, ip_map, host_map)
    ov_hm = enrich_with_assets(ov_hm, ip_map, host_map)

    # Derive scan_date from first non-empty timestamp, fallback to today
    ts = ov_full.get("timestamp")
    scan_date = None
    if ts is not None and len(ts):
        val = str(next((t for t in ts if str(t).strip()), "")
        )
        scan_date = val[:10] if val else None
    if not scan_date:
        scan_date = datetime.now().strftime("%Y-%m-%d")

    # YYMMDD string for sheet naming
    scan_yymmdd = datetime.strptime(scan_date, "%Y-%m-%d").strftime("%y%m%d")
    details_name = f"OpenVAS_{scan_yymmdd}"
    summary_name = "OpenVAS_Summary"

    # --- Accumulated Summary logic ---
    existing_detail_sheets = {}
    existing_summary = None
    existing_meta = None
    if output_xlsx.exists():
        try:
            all_existing = pd.read_excel(output_xlsx, sheet_name=None)
            existing_summary = all_existing.get(summary_name)
            existing_meta = all_existing.get("_meta")
            # Keep only prior details named OpenVAS_YYMMDD (not the one we're about to write)
            for nm, df in all_existing.items():
                if nm == summary_name or nm == "_meta":
                    continue
                if nm == details_name:
                    continue
                if nm.startswith("OpenVAS_"):
                    existing_detail_sheets[nm] = df
        except Exception:
            existing_detail_sheets = {}
            existing_summary = None
            existing_meta = None

    # Build current NOW counts (H/M only) per host
    def pick_display(r):
        nm = str(r.get("asset_name") or "").strip()
        if nm: return nm
        hn = str(r.get("hostname") or "").strip()
        if hn: return hn
        ip = str(r.get("ip") or "").strip()
        return ip if ip else "unknown"

    ov_hm_view = ov_hm.copy()
    ov_hm_view["host"] = ov_hm_view.apply(pick_display, axis=1)
    sev_norm_hm = ov_hm_view["severity"].astype(str).str.strip().str.lower()
    # Counts per host
    hm_counts = ov_hm_view.groupby(["host", sev_norm_hm]).size().unstack(fill_value=0)
    # Ensure both columns exist
    for col in ("high", "medium"):
        if col not in hm_counts.columns:
            hm_counts[col] = 0
    hm_counts = hm_counts.reset_index().rename(columns={"high": "high_count_now", "medium": "medium_count_now"})
    hm_counts["totals_now"] = hm_counts["high_count_now"] + hm_counts["medium_count_now"]

    # Attach asset fields (first per host from enriched FULL df)
    full_view = ov_full.copy()
    full_view["host"] = full_view.apply(pick_display, axis=1)
    asset_first = full_view.groupby("host").agg({
        "asset_status": "first",
        "asset_location": "first",
        "asset_source": "first"
    }).reset_index()

    now_df = hm_counts.merge(asset_first, on="host", how="left")

    # Column name for this run's H+M totals in the summary
    # If this is the first run (no existing summary), use baseline-YYMMDD-totals; otherwise scan-YYMMDD-totals
    scan_col = None  # set after we know whether a prior summary exists

    # Merge with existing summary (if any), preserving prior scan-* columns and updating NOW columns
    if existing_summary is not None and not existing_summary.empty:
        # Determine if we've already processed this scan date
        baseline_col = f"baseline-{scan_yymmdd}-totals"
        scan_already = f"scan-{scan_yymmdd}-totals"
        cols_exist = set(map(str, existing_summary.columns))
        already_baseline = baseline_col in cols_exist
        already_scan = scan_already in cols_exist

        # On subsequent runs we normally add scan-YYMMDD-totals, but
        # if this data-stamp was already used as the baseline or scan, don't add another column.
        add_scan_col = not (already_baseline or already_scan)
        if add_scan_col:
            scan_col = scan_already
        else:
            scan_col = None  # we'll only refresh the NOW + asset_* fields

        # Track all prior per-scan columns we must preserve
        prior_scan_cols = [
            c for c in existing_summary.columns
            if str(c).startswith("scan-") or str(c).startswith("baseline-")
        ]

        # Start with existing summary, but drop old NOW columns only (NOT asset_* fields)
        drop_cols = {"high_count_now", "medium_count_now", "totals_now"}
        keep_cols = [c for c in existing_summary.columns if c not in drop_cols]
        base = existing_summary[keep_cols]

        # Rename asset columns in now_df so we can coalesce after merge
        now_df = now_df.rename(columns={
            "asset_status": "now_asset_status",
            "asset_location": "now_asset_location",
            "asset_source": "now_asset_source",
        })

        # Add this run's scan column if needed
        if add_scan_col:
            now_df[scan_col] = now_df["totals_now"]

        # Merge base with new snapshot
        merged = base.merge(now_df, on="host", how="outer")

        # Coalesce asset fields: prefer current, else keep prior
        import numpy as _np
        for col in ("status", "location", "source"):
            prior = f"asset_{col}"
            now = f"now_asset_{col}"
            if prior in merged.columns and now in merged.columns:
                merged[prior] = merged[now].where(merged[now].astype(str).str.strip() != "", merged[prior])
                merged = merged.drop(columns=[now])

        # Ensure prior scan columns are retained (outer merge may introduce NaN); re-add any missing
        for c in prior_scan_cols:
            if c not in merged.columns:
                merged[c] = existing_summary.get(c)

        # Optional: note to console if we skipped adding a scan column
        if not add_scan_col:
            print(f"[openvas-25] Note: scan date {scan_yymmdd} already processed (baseline or scan column exists); refreshed NOW and asset fields only.")

        summary_out = merged
    else:
        # First run for this output: create baseline with the current scan column
        scan_col = f"baseline-{scan_yymmdd}-totals"
        now_df[scan_col] = now_df["totals_now"]
        summary_out = now_df

    # --- Post-merge normalization and status rules ---
    # Ensure NOW counts are numeric and missing values become 0
    for c in ("high_count_now", "medium_count_now", "totals_now"):
        if c in summary_out.columns:
            summary_out[c] = pd.to_numeric(summary_out[c], errors="coerce").fillna(0).astype(int)
        else:
            summary_out[c] = 0

    # Ensure columns exist and are string-typed
    for col in ("asset_status", "asset_location", "asset_source"):
        if col not in summary_out.columns:
            summary_out[col] = ""
        summary_out[col] = summary_out[col].fillna("")

    # Only mark "not in scan" when there are zero findings NOW AND there was no prior status at all
    no_now_mask = (summary_out["totals_now"] == 0) & (summary_out["asset_status"].astype(str).str.strip() == "")
    summary_out.loc[no_now_mask, "asset_status"] = "not in scan"
    # Do NOT default asset_source to openvas; preserve whatever was in the tracker/previous run

    # --- Refresh asset fields from Asset Tracker when we have better truth ---
    # If prior runs defaulted asset_source to openvas for hosts that actually exist
    # in the tracker, prefer tracker data for source/location (and status if blank).
    refreshed_rows = []
    for idx, r in summary_out.iterrows():
        host_disp = str(r.get("host") or "").strip()
        prior_source = str(r.get("asset_source") or "").strip().lower()
        prior_loc = str(r.get("asset_location") or "").strip()
        prior_status = str(r.get("asset_status") or "").strip()

        match = lookup_asset_for_display(host_disp, ip_map, host_map)
        if match is None:
            refreshed_rows.append(r)
            continue

        # Only upgrade fields if we have better (non-empty and not the generic 'openvas')
        new_source = str(match.get("source", "")).strip()
        new_loc = str(match.get("location", "")).strip()
        new_status = str(match.get("status", "")).strip()

        if (prior_source == "" or prior_source == "openvas") and new_source:
            r["asset_source"] = new_source
        if (prior_loc == "") and new_loc:
            r["asset_location"] = new_loc
        # Do not overwrite explicit statuses like PoweredOn/PoweredOff/not in scan.
        # Only fill status if it was blank.
        if prior_status == "" and new_status:
            r["asset_status"] = new_status

        refreshed_rows.append(r)

    if refreshed_rows:
        summary_out = pd.DataFrame(refreshed_rows, columns=summary_out.columns)

    # Prune legacy/undesired columns from older versions
    # Do NOT drop new scan-YYMMDD-totals columns; only remove legacy columns
    drop_patterns = [
        r"^Totals-",               # legacy prefix from earlier builds
        r"^high$",
        r"^medium$",
        r"^total_findings$",
        r"^high_count$",
        r"^medium_count$",
    ]
    to_drop = []
    for col in summary_out.columns:
        for pat in drop_patterns:
            if re.match(pat, str(col)):
                to_drop.append(col)
                break
    if to_drop:
        summary_out = summary_out.drop(columns=to_drop)

    baseline_cols = sorted([c for c in summary_out.columns if str(c).startswith("baseline-")])
    scan_cols = sorted([c for c in summary_out.columns if str(c).startswith("scan-")])
    allowed_core = [
        "host", "high_count_now", "medium_count_now", "totals_now",
        "asset_status", "asset_location", "asset_source",
    ]
    ordered_cols = allowed_core + baseline_cols + scan_cols
    # Keep only allowed columns in that order
    summary_out = summary_out[[c for c in ordered_cols if c in summary_out.columns]]

    # Sort view for usability: current NOW columns first, then baseline totals (if present), then host
    sort_keys = ["totals_now", "high_count_now", "medium_count_now"]
    sort_asc = [False, False, False]
    # Add baseline-YYMMDD-totals as an additional tiebreaker if it exists
    if baseline_cols:
        sort_keys.append(baseline_cols[0])  # there should only be one baseline column
        sort_asc.append(False)
    sort_keys.append("host")
    sort_asc.append(True)

    summary_out = summary_out.sort_values(by=sort_keys, ascending=sort_asc).reset_index(drop=True)

    # Update meta
    new_meta = pd.DataFrame([{
        "openvas_csv_path": str(latest),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "scan_date": scan_date,
        "details_sheet": details_name
    }])
    if existing_meta is not None and not existing_meta.empty:
        meta_out = pd.concat([existing_meta, new_meta], ignore_index=True)
    else:
        meta_out = new_meta

    # --- Write workbook: Summary first, then preserve prior detail sheets, then this run's details, then _meta
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as w:
        # 1) Summary first
        summary_out.to_excel(w, sheet_name=summary_name, index=False)

        # 2) Preserve prior detail sheets only (OpenVAS_YYMMDD)
        for nm in sorted(existing_detail_sheets.keys()):
            existing_detail_sheets[nm].to_excel(w, sheet_name=nm, index=False)

        # 3) New details sheet for this scan
        ov_hm.to_excel(w, sheet_name=details_name, index=False)

        # 4) Meta
        meta_out.to_excel(w, sheet_name="_meta", index=False)

        # --- Auto-fit widths on Summary
        ws = w.book[summary_name]
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row,
                                                         min_col=1, max_col=ws.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

        # Auto-fit widths on current details sheet
        ws_d = w.book[details_name]
        for col_idx, col_cells in enumerate(ws_d.iter_cols(min_row=1, max_row=ws_d.max_row,
                                                           min_col=1, max_col=ws_d.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws_d.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

    print(f"[openvas-25] Wrote: {output_xlsx} (source: {latest}) details={details_name} summary={summary_name}")

# -------------------------------
# CLI
# -------------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Build OpenVAS_25 workbook: Summary tab (OpenVAS_Summary) accumulates H+M totals per host across scans, with current high/medium/totals and a per-scan H+M column (first run: baseline-YYMMDD-totals; subsequent: scan-YYMMDD-totals). Details tab is High/Medium only and date-stamped as OpenVAS_YYMMDD. _meta is appended each run."
    )
    ap.add_argument("--asset-tracker", required=True, help="Path to Asset_Tracker.xlsx (must contain 'asset_tracker' sheet).")
    ap.add_argument("--openvas-root", required=True, help="Root folder containing detailedresults.csv files (searches recursively).")
    ap.add_argument("--output", required=True, help="Output XLSX path.")
    args = ap.parse_args()

    try:
        build_openvas_book(
            Path(args.asset_tracker),
            Path(args.openvas_root),
            Path(args.output)
        )
    except FileNotFoundError as e:
        print(f"[openvas-25] ERROR: {e}")
        sys.exit(2)
    except Exception as e:
        # Unexpected error — show class and message, but avoid a huge stack unless -v is desired
        print(f"[openvas-25] Unexpected error: {e.__class__.__name__}: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
