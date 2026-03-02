#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Infrastructure Asset + Vulnerability Tracker

Rules enforced:
- Normalize names by LOWERCASING ONLY (no suffix stripping, no dot trimming).
- vCenter status sync:
  * Match by IP, or by name/dns_name vs vCenter Name/Hostname (all lowercased).
  * If matched: set status = vCenter PowerState; set location = vc_vmhost (if present);
    append vc_notes to notes.
  * If not matched and asset is virtual: status = "not in vcenter".
- OpenVAS parsing is strict:
  * Use only the 'IP' and 'Hostname' columns (Hostname lowercased).
  * No IP-range filtering.
  * Collect CPEs (cpe:/o: preferred, else cpe:/a:) from same-row text fields.
- OpenVAS effects:
  * If an asset is "not in vcenter" but appears in OpenVAS (by IP or hostname), set status = "online".
  * Add net-new IPs from OpenVAS as rows with status="found in scan", source="openvas".
  * Add OS guess "Found OS: <cpe>" ONLY to rows with status="found in scan".
"""

import argparse
import re
import ipaddress
from datetime import datetime
from pathlib import Path
from typing import Optional
from shutil import copy2
import pandas as pd

# Optional YAML support for --config
try:
    import yaml  # pip install pyyaml
except ImportError:
    yaml = None


# ----------------------------
# Constants (to avoid magic strings / Sonar S1192)
# ----------------------------
LIT_IP_ADDRESS = "IP Address"
LIT_IP_ADDRESS_ALT1 = "IP Address.1"
STATUS_NOT_IN_VCENTER = "not in vcenter"
STATUS_FOUND_IN_SCAN = "found in scan"
NOTE_PRESUME_OFFLINE = "machine presumed offline -- unless OpenVAS says otherwise"

# Additional constants
STATUS_ONLINE = "online"
STATUS_PHYSICAL = "physical"
NOTE_FOUND_VCENTER_MISS = "machine was not found in vcenter but is in OpenVAS results"
SOURCE_OPENVAS = "openvas"
SOURCE_VIRTUAL = "virtual"

# ----------------------------
# Helpers
# ----------------------------

def strip_bom_headers(cols):
    return [c.encode("utf-8").decode("utf-8-sig") for c in cols]

IPV4_RE = re.compile(r"\b((?:\d{1,3}\.){3}\d{1,3})\b")

def first_ipv4(s):
    if not s:
        return None
    m = IPV4_RE.search(str(s))
    return m.group(1) if m else None

def extract_all_ipv4s(s):
    return IPV4_RE.findall(str(s or ""))

def lower_or_blank(x):
    return str(x).lower().strip() if (x is not None and str(x).strip() != "") else ""

def parse_vc_ts_from_filename(name):
    m = re.search(r"vsphere_vmlist_(\d{8})-(\d{6})", name)
    if m:
        ymd, hms = m.groups()
        return datetime(int(ymd[:4]), int(ymd[4:6]), int(ymd[6:]),
                        int(hms[:2]), int(hms[2:4]), int(hms[4:]))
    return None

def newest_detailedresults_csv(root: Path):
    files = list(root.rglob("detailedresults.csv"))
    if not files:
        raise FileNotFoundError(f"No detailedresults.csv found under {root}")
    files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0]

def scrub(df: pd.DataFrame):
    df = df.copy()
    # standardize empties to None, then back to NaN via pandas if needed
    for c in df.columns:
        df[c] = df[c].replace({"none": None, "None": None, "": None})
    return df

def coerce_ip_list(x):
    if isinstance(x, list):
        return [str(ip).strip() for ip in x if ip]
    if x is None or pd.isna(x):
        return []
    return extract_all_ipv4s(str(x))

def ip_valid(ip_str: str) -> bool:
    try:
        ip = ipaddress.ip_address(ip_str)
        return ip.version == 4
    except Exception:
        return False

# ----------------------------
# Loaders
# ----------------------------

def load_physical(path: Path):
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    df.columns = [c.strip() for c in strip_bom_headers(df.columns)]
    ip1_col = LIT_IP_ADDRESS_ALT1 if LIT_IP_ADDRESS_ALT1 in df.columns else LIT_IP_ADDRESS
    phys = pd.DataFrame({
        "name": df.get("Device - Model"),
        "ip_address": df[ip1_col].apply(first_ipv4) if ip1_col in df.columns else None,
        "data_classification": df.get("Classification"),
        "type_user": df.get("Type"),
        "purpose": df.get("Function"),
        "dns_name": None,
        "location": df.get("Location"),
        "status": "physical",
        "svc_tag": df.get("SVC Tag"),
        "source": "physical",
        "notes": None
    })
    return scrub(phys), df

def load_virtual(path: Path):
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    df.columns = [c.strip() for c in strip_bom_headers(df.columns)]
    virt = pd.DataFrame({
        "name": df.get("Name"),
        "ip_address": df.get(LIT_IP_ADDRESS).apply(first_ipv4) if LIT_IP_ADDRESS in df.columns else None,
        "data_classification": df.get("Data Classification"),
        "type_user": df.get("Type / User"),
        "purpose": df.get("Purpose"),
        "dns_name": df.get("DNS Name"),
        "location": SOURCE_VIRTUAL,
        "status": df.get("PowerStatus_CVI_25"),
        "svc_tag": None,
        "source": SOURCE_VIRTUAL,
        "notes": None
    })
    return scrub(virt), df

def load_vcenter(path: Path):
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    df.columns = [c.strip() for c in strip_bom_headers(df.columns)]

    name_col = next((c for c in ("Name", "VMName", "name") if c in df.columns), None)
    host_name_col = next((c for c in ("Hostname", "DNS Name", "Guest Hostname", "guest_hostname", "host_name") if c in df.columns), None)
    pcol = next((c for c in ("PowerState", "powerstate") if c in df.columns), None)
    vmhost_col = next((c for c in ("VMHost", "Host", "VM Host", "ESX Host") if c in df.columns), None)
    notes_col = next((c for c in ("Notes", "Annotation", "notes") if c in df.columns), None)
    multi_ip_col = next((c for c in ("IPAddresses", "IP Addresses", "Guest IP Address", "Guest IP Addresses") if c in df.columns), None)
    ip_col_single = next((c for c in ("IP Address", "IP", "ip_address") if c in df.columns), None)

    vc_name = df[name_col].map(lower_or_blank) if name_col else pd.Series([""] * len(df))
    vc_host = df[host_name_col].map(lower_or_blank) if host_name_col else pd.Series([""] * len(df))

    if multi_ip_col:
        ip_lists = df[multi_ip_col].apply(extract_all_ipv4s)
    elif ip_col_single:
        ip_lists = df[ip_col_single].apply(extract_all_ipv4s)
    else:
        ip_lists = pd.Series([[]] * len(df))

    ip_primary = ip_lists.apply(lambda lst: lst[0] if lst else None)
    ts = parse_vc_ts_from_filename(path.name) or datetime.fromtimestamp(path.stat().st_mtime)

    slim = pd.DataFrame({
        "vc_name": vc_name,
        "vc_host": vc_host,
        "vc_powerstate": df[pcol] if pcol else None,
        "vc_vmhost": df[vmhost_col] if vmhost_col else None,
        "vc_notes": df[notes_col] if notes_col else None,
        "ip_addresses": ip_lists,
        "ip_address": ip_primary,
        "vcenter_inventory_timestamp": ts
    })
    return scrub(slim), df

# ----------------------------
# OpenVAS (STRICT: IP + HOSTNAME + CPE extraction)
# ----------------------------

CPE_RE = re.compile(r"(cpe:/[aoh]:[^,\s\"']+)", re.IGNORECASE)

def parse_openvas(root: Path):
    """
    Returns (ip_set, hostname_set, cpe_map_by_ip, csv_path)
    - ip_set: IPs from IP column
    - hostname_set: normalized(lowercase) hostnames from Hostname column
    - cpe_map_by_ip: dict ip -> [cpe strings]
    """
    f = newest_detailedresults_csv(root)
    df = pd.read_csv(f, dtype=str, encoding="utf-8-sig").fillna("")
    # normalize column detection
    def find_col(target_lower):
        for c in df.columns:
            if c.strip().lower() == target_lower:
                return c
        return None

    ip_col = find_col("ip")
    host_col = find_col("hostname")
    cpe_cols = [c for c in df.columns if c in {
        "Affected Software/OS",
        "Product Detection Result",
        "Specific Result",
        "Summary"
    }]

    if ip_col is None or host_col is None:
        raise RuntimeError("OpenVAS CSV must contain 'IP' and 'Hostname' columns.")

    ip_set = set()
    host_set = set()
    cpe_map = {}

    for _, row in df.iterrows():
        ip_raw = str(row.get(ip_col, "")).strip()
        host_raw = lower_or_blank(row.get(host_col, ""))

        if host_raw:
            host_set.add(host_raw)

        if ip_valid(ip_raw):
            ip_key = ip_raw.lower()
            ip_set.add(ip_key)

            # collect CPEs from the same row
            cpes_here = []
            for tc in cpe_cols:
                text = str(row.get(tc, "") or "")
                for m in CPE_RE.finditer(text):
                    cpe = m.group(1).lower()
                    if cpe not in cpes_here:
                        cpes_here.append(cpe)
            if cpes_here:
                cpe_map.setdefault(ip_key, [])
                for cpe in cpes_here:
                    if cpe not in cpe_map[ip_key]:
                        cpe_map[ip_key].append(cpe)

    return ip_set, host_set, cpe_map, f

# ----------------------------
# Builder
# ----------------------------

def build_asset_tracker(phys_path, virt_path, vc_path, openvas_root, output):
    # Load
    phys, phys_tab = load_physical(phys_path)
    virt, virt_tab = load_virtual(virt_path)
    vc, vc_tab = load_vcenter(vc_path)
    ov_ip_set, ov_host_set, ov_cpe_map, _ov_file = parse_openvas(openvas_root)  # path kept for provenance if needed

    # Provenance: capture the exact OpenVAS CSV path and a single vCenter inventory timestamp
    ov_file_path = str(_ov_file)
    vc_ts = None
    if "vcenter_inventory_timestamp" in vc.columns:
        non_null_ts = vc["vcenter_inventory_timestamp"].dropna()
        if len(non_null_ts):
            vc_ts = non_null_ts.iloc[0]

    # Base asset set
    asset = pd.concat([phys, virt], ignore_index=True)

    # Normalize names by LOWERCASE ONLY
    asset["name"] = asset["name"].map(lower_or_blank)
    asset["dns_name"] = asset["dns_name"].map(lower_or_blank)

    # vCenter matching indexes
    # Build IP->vc rows
    vc_ip_map = {}
    for _, r in vc.iterrows():
        for ip in coerce_ip_list(r.get("ip_addresses")):
            vc_ip_map.setdefault(ip.lower(), []).append(r)

    # Build name/dns -> vc rows
    vc_name_map = {}
    def add_key(d, k, row):
        if not k:
            return
        d.setdefault(k, []).append(row)

    for _, r in vc.iterrows():
        add_key(vc_name_map, r.get("vc_name", ""), r)
        add_key(vc_name_map, r.get("vc_host", ""), r)

    # 1) vCenter status & enrichment
    # match by IP first, then by name/dns versus vCenter name/host (unique match only for name/dns)
    def match_vcenter_row(a_row):
        # by IP
        ip = (a_row.get("ip_address") or "").lower()
        if ip and ip in vc_ip_map:
            return vc_ip_map[ip][0]

        # by normalized name/dns_name — only accept a unique match
        nm = (a_row.get("name") or "")
        dn = (a_row.get("dns_name") or "")
        candidates = []
        for key in (nm, dn):
            if key and key in vc_name_map:
                candidates.extend(vc_name_map[key])
        if candidates:
            uniq = { (r.get("vc_name",""), r.get("vc_host",""), r.get("vc_vmhost","")) : r for r in candidates }
            return list(uniq.values())[0] if len(uniq) == 1 else None
        return None

    # work on a copy to avoid pandas SettingWithCopy warnings
    j = asset.copy()

    # enrich per row
    vc_status_col = "vc_powerstate"
    vc_vmhost_col = "vc_vmhost"
    vc_notes_col = "vc_notes"

    new_status = []
    new_location = []
    new_notes = []

    for _, row in j.iterrows():
        cur_status = row.get("status")
        cur_location = row.get("location")
        cur_notes = row.get("notes")

        hit = match_vcenter_row(row)
        if hit is not None:
            # adopt vCenter powerstate
            s = hit.get(vc_status_col) or cur_status
            # set location to VM host only for virtual assets
            if (row.get("source") == SOURCE_VIRTUAL) and hit.get(vc_vmhost_col):
                loc = hit.get(vc_vmhost_col)
            else:
                loc = cur_location
            # merge notes
            vcn = hit.get(vc_notes_col)
            if vcn and vcn != cur_notes and vcn not in (cur_notes or ""):
                notes = (cur_notes + "; " + vcn) if cur_notes else vcn
            else:
                notes = cur_notes
        else:
            # no match in vCenter
            if (row.get("source") == SOURCE_VIRTUAL):
                s = STATUS_NOT_IN_VCENTER
                extra_note = NOTE_PRESUME_OFFLINE
                if cur_notes:
                    notes = f"{cur_notes}; {extra_note}" if extra_note not in cur_notes else cur_notes
                else:
                    notes = extra_note
            else:
                s = cur_status
                notes = cur_notes
            loc = cur_location

        new_status.append(s)
        new_location.append(loc)
        new_notes.append(notes)

    j["status"] = new_status
    j["location"] = new_location
    j["notes"] = new_notes

    # 2) OpenVAS: bring online if "not in vcenter" but seen in OpenVAS (by IP or hostname)
    # prepare fast sets
    known_ips = {ip.lower() for ip in j["ip_address"].dropna().astype(str)} #lower added for IPv6
    ov_names = ov_host_set  # already lowercased
    ov_ips = ov_ip_set

    def seen_in_openvas(a_row):
        ip = (a_row.get("ip_address") or "").lower()
        if ip and ip in ov_ips:
            return True
        nm = (a_row.get("name") or "")
        dn = (a_row.get("dns_name") or "")
        return (nm in ov_names) or (dn in ov_names)

    mask_not_in_vc = j["status"].map(lambda s: (s or "").lower() == STATUS_NOT_IN_VCENTER)
    mask_seen = mask_not_in_vc & j.apply(seen_in_openvas, axis=1)

    # Set status to online for those seen in OpenVAS
    j.loc[mask_seen, "status"] = STATUS_ONLINE

    # Update notes: remove the presumption note; add explicit OpenVAS confirmation
    old_note = NOTE_PRESUME_OFFLINE
    new_note = NOTE_FOUND_VCENTER_MISS
    for idx in j.index[mask_seen]:
        cur_notes = j.at[idx, "notes"] or ""
        # remove old note variants
        cleaned = cur_notes.replace(f"; {old_note}", "").replace(old_note, "").strip()
        # append new note if not already present
        if new_note not in cleaned:
            cleaned = (cleaned + "; " + new_note).strip("; ").strip()
        j.at[idx, "notes"] = cleaned

    # 3) Add net-new IPs from OpenVAS (not present by IP in PI/VI)
    add_rows = []
    for ip in sorted(ov_ips):
        if ip and ip not in known_ips:
            add_rows.append({
                "name": "",
                "ip_address": ip,
                "data_classification": None,
                "type_user": None,
                "purpose": None,
                "dns_name": "",
                "location": None,
                "status": STATUS_FOUND_IN_SCAN,
                "svc_tag": None,
                "source": SOURCE_OPENVAS,
                "notes": None
            })
    if add_rows:
        add = pd.DataFrame(add_rows)
        j = pd.concat([j, add], ignore_index=True, sort=False)

    # 4) Append OS guess from CPE into notes ONLY for 'found in scan'
    def enrich_notes_with_os(row):
        status = (row.get("status") or "").strip().lower()
        if status != STATUS_FOUND_IN_SCAN:
            return row.get("notes")
        existing = row.get("notes") or ""
        ip = str(row.get("ip_address") or "").lower()
        cpes = ov_cpe_map.get(ip, [])
        if cpes:
            # prefer OS first, then app
            os_guess = next((c for c in cpes if c.startswith("cpe:/o:")), None)
            if not os_guess:
                os_guess = next((c for c in cpes if c.startswith("cpe:/a:")), None)
            if os_guess:
                return (existing + "; " if existing else "") + f"Found OS: {os_guess}"
        return existing

    j["notes"] = j.apply(enrich_notes_with_os, axis=1)

    # Helper sheet: flatten vCenter IPs (one row per IP)
    ip_rows = []
    for _, r in vc.iterrows():
        for ip in coerce_ip_list(r.get("ip_addresses")):
            if not ip_valid(ip):
                continue
            ip_rows.append({
                "name": r.get("vc_name"),
                "dns_name": r.get("vc_host"),
                "ip_address": ip,
                "vc_powerstate": r.get("vc_powerstate"),
                "vc_vmhost": r.get("vc_vmhost"),
                "vcenter_inventory_timestamp": r.get("vcenter_inventory_timestamp")
            })
    asset_ips = pd.DataFrame(ip_rows) if ip_rows else pd.DataFrame(
        columns=["name","dns_name","ip_address","vc_powerstate","vc_vmhost","vcenter_inventory_timestamp"]
    )

    # Meta sheet with provenance and build time
    meta = pd.DataFrame([{
        "openvas_csv_path": ov_file_path,
        "vcenter_inventory_timestamp": (None if vc_ts is None else str(vc_ts)),
        "generated_at": datetime.now().isoformat(timespec="seconds")
    }])

    # Final sort: asset tracker by numeric IP address (then name as stable tiebreak)
    def _ip_sort_key(val):
        try:
            parts = [int(p) for p in str(val).split(".")]
            while len(parts) < 4:
                parts.append(0)
            return tuple(parts[:4])
        except Exception:
            return (999, 999, 999, 999)

    j = j.copy()
    j["_ip_sort"] = j["ip_address"].map(_ip_sort_key)
    j = j.sort_values(by=["_ip_sort", "name"], kind="mergesort").drop(columns=["_ip_sort"])

    # Canonicalize Status values to improve Excel filter reliability
    j["status"] = j["status"].astype(str).str.strip()
    j["status"] = j["status"].replace({
        r"(?i)^powered\s*off$": "PoweredOff",
        r"(?i)^powered\s*on$": "PoweredOn",
        r"(?i)^not\s*in\s*v\s*center$": STATUS_NOT_IN_VCENTER,
        r"(?i)^not\s*in\s*vcenter$": STATUS_NOT_IN_VCENTER,
        r"(?i)^found\s*in\s*scan$": STATUS_FOUND_IN_SCAN,
        r"(?i)^online$": STATUS_ONLINE,
        r"(?i)^physical$": STATUS_PHYSICAL
    }, regex=True)

    # Precompute allowed statuses for Excel filter (exclude 'PoweredOff' and 'not in vcenter')
    unique_statuses = [str(s) for s in j["status"].dropna().unique().tolist()]
    blocked = {"poweredoff", STATUS_NOT_IN_VCENTER}
    allowed_statuses = [s for s in unique_statuses if s.strip().lower() not in blocked]
    status_col_index_0b = j.columns.get_loc("status")  # zero-based for openpyxl filter colId

    # Write workbook (asset_tracker first) and apply Excel autofilter to exclude 'PoweredOff'
    with pd.ExcelWriter(output, engine="openpyxl") as w:
        j.to_excel(w, sheet_name="asset_tracker", index=False)
        phys_tab.to_excel(w, sheet_name="physical_inventory_2024", index=False)
        virt_tab.to_excel(w, sheet_name="virtual_inventory_2024", index=False)
        vc_tab.to_excel(w, sheet_name="vCenter_View", index=False)
        asset_ips.to_excel(w, sheet_name="asset_ips", index=False)
        meta.to_excel(w, sheet_name="_meta", index=False)

        # Apply AutoFilter on Status column to exclude 'PoweredOff' and 'not in vcenter'
        ws = w.book["asset_tracker"]
        # Define filter range to all used data (A1:...)
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(ws.max_column)
        last_row = ws.max_row
        # Reset any stale autofilter first (Excel/Mac quirk), then set the new range
        ws.auto_filter.ref = None
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # Auto-fit column widths on asset_tracker (approximate by content length)
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row,
                                                         min_col=1, max_col=ws.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            # reasonable bounds
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

        # Select only allowed statuses (everything except 'PoweredOff' and 'not in vcenter')
        if allowed_statuses:
            ws.auto_filter.add_filter_column(status_col_index_0b, allowed_statuses, blank=True)
            # Workaround: explicitly hide rows whose Status is 'PoweredOff' or 'not in vcenter' so the initial view is filtered
            blocked_vals = {"poweredoff", STATUS_NOT_IN_VCENTER}
            status_col_1b = status_col_index_0b + 1
            for r in range(2, last_row + 1):
                val = ws.cell(row=r, column=status_col_1b).value
                if isinstance(val, str) and val.strip().lower() in blocked_vals:
                    ws.row_dimensions[r].hidden = True

    print(f"[build-asset-tracker] Wrote: {output}")
    print(f"[build-asset-tracker] OpenVAS file: {ov_file_path}")


# ----------------------------
# Updater (incremental)
# ----------------------------

def update_asset_tracker(baseline_xlsx: Path, vcenter_path: Optional[Path], openvas_root: Optional[Path], output: Optional[Path]):
    """
    Incrementally update an existing Asset_Tracker.xlsx with new vCenter and/or OpenVAS.
    Baseline defaults to ./outputs/Asset_Tracker.xlsx. If output is None, overwrite baseline
    after backing it up to Asset_Tracker_YYYYMMDD-HHMMSS.xlsx.
    Behavior:
      - vCenter provided: IP-first match (then unique name/dns). vCenter powerstate wins; set
        location=vc_vmhost for virtuals; merge vc_notes; if virtual not found and not already in
        {not in vcenter, online}, set status="not in vcenter" and append presumption note.
      - OpenVAS provided: flip not-in-vcenter -> online when seen by IP or hostname; add net-new
        IPs as status="found in scan" with OS guess in notes.
      - Recreate filters on asset_tracker (hide PoweredOff + not in vcenter) and auto-fit columns.
      - If vCenter refreshed, replace vCenter_View and rebuild asset_ips.
    """
    # Resolve defaults and backup if overwriting baseline
    if output is None:
        output = baseline_xlsx  # in-place
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = baseline_xlsx.with_name(f"{baseline_xlsx.stem}_{ts}{baseline_xlsx.suffix}")
        copy2(baseline_xlsx, backup)
        print(f"[update-asset-tracker] Backed up baseline to: {backup}")

    # --- Load baseline workbook
    sheets = pd.read_excel(baseline_xlsx, sheet_name=None, dtype=str)
    sheets = {k: v.fillna("") for k, v in sheets.items()}
    if "asset_tracker" not in sheets:
        raise RuntimeError(f"Baseline {baseline_xlsx} has no 'asset_tracker' sheet")
    j = sheets["asset_tracker"].copy()

    # Normalize name/dns; ensure ip column exists
    if "name" in j.columns: j["name"] = j["name"].map(lower_or_blank)
    else: j["name"] = ""
    if "dns_name" in j.columns: j["dns_name"] = j["dns_name"].map(lower_or_blank)
    else: j["dns_name"] = ""
    if "ip_address" not in j.columns: j["ip_address"] = ""

    # --- Optional vCenter refresh (vCenter status wins when matched)
    vc = None
    vc_tab = None
    vc_ts = None
    if vcenter_path:
        vc, vc_tab = load_vcenter(vcenter_path)
        non_null_ts = vc.get("vcenter_inventory_timestamp", pd.Series([])).dropna()
        if len(non_null_ts): vc_ts = non_null_ts.iloc[0]

        # Maps
        vc_ip_map: dict[str, list[pd.Series]] = {}
        for _, r in vc.iterrows():
            for ip in coerce_ip_list(r.get("ip_addresses")):
                vc_ip_map.setdefault(str(ip).lower(), []).append(r)

        vc_name_map: dict[str, list[pd.Series]] = {}
        def add_key(d, k, row):
            if k: d.setdefault(k, []).append(row)
        for _, r in vc.iterrows():
            add_key(vc_name_map, r.get("vc_name",""), r)
            add_key(vc_name_map, r.get("vc_host",""), r)

        def match_vcenter_row(a_row):
            ip = (a_row.get("ip_address") or "").lower()
            if ip and ip in vc_ip_map:
                return vc_ip_map[ip][0]
            nm = (a_row.get("name") or "")
            dn = (a_row.get("dns_name") or "")
            candidates = []
            for key in (nm, dn):
                if key and key in vc_name_map:
                    candidates.extend(vc_name_map[key])
            if candidates:
                uniq = {(r.get("vc_name",""), r.get("vc_host",""), r.get("vc_vmhost","")): r for r in candidates}
                return list(uniq.values())[0] if len(uniq) == 1 else None
            return None

        vc_status_col = "vc_powerstate"
        vc_vmhost_col = "vc_vmhost"
        vc_notes_col  = "vc_notes"

        new_status, new_location, new_notes = [], [], []
        for _, row in j.iterrows():
            cur_status   = row.get("status")
            cur_location = row.get("location")
            cur_notes    = row.get("notes") or ""

            hit = match_vcenter_row(row)
            if hit is not None:
                # vCenter wins on status
                s = hit.get(vc_status_col) or cur_status
                # vmhost only for virtual
                if (row.get("source") == SOURCE_VIRTUAL) and hit.get(vc_vmhost_col):
                    loc = hit.get(vc_vmhost_col)
                else:
                    loc = cur_location
                # notes: remove presumption, then add vc_notes if not present
                notes = cur_notes
                if NOTE_PRESUME_OFFLINE in notes:
                    notes = notes.replace(f"; {NOTE_PRESUME_OFFLINE}", "").replace(NOTE_PRESUME_OFFLINE, "").strip()
                vcn = hit.get(vc_notes_col)
                if vcn and vcn not in notes:
                    notes = (notes + "; " + vcn) if notes else vcn
            else:
                # Not found in vCenter: flip to 'not in vcenter' for virtuals unless already special states
                if (row.get("source") == SOURCE_VIRTUAL):
                    cur_lower = (cur_status or "").lower()
                    if cur_lower not in {STATUS_NOT_IN_VCENTER, STATUS_ONLINE}:
                        s = STATUS_NOT_IN_VCENTER
                        notes = cur_notes
                        if NOTE_PRESUME_OFFLINE not in notes:
                            notes = (notes + "; " + NOTE_PRESUME_OFFLINE) if notes else NOTE_PRESUME_OFFLINE
                    else:
                        s = cur_status
                        notes = cur_notes
                else:
                    s = cur_status
                    notes = cur_notes
                loc = cur_location

            new_status.append(s)
            new_location.append(loc)
            new_notes.append(notes)

        j["status"]   = new_status
        j["location"] = new_location
        j["notes"]    = new_notes

    # --- Optional OpenVAS refresh (flip + add net-new IPs)
    ov_file_path = None
    if openvas_root:
        ov_ip_set, ov_host_set, ov_cpe_map, _ov_file = parse_openvas(openvas_root)
        ov_file_path = str(_ov_file)

        known_ips = {ip.lower() for ip in j["ip_address"].dropna().astype(str)}

        def seen_in_openvas(a_row):
            ip = (a_row.get("ip_address") or "").lower()
            if ip and ip in ov_ip_set:
                return True
            nm = (a_row.get("name") or "")
            dn = (a_row.get("dns_name") or "")
            return (nm in ov_host_set) or (dn in ov_host_set)

        # Flip 'not in vcenter' -> 'online' when seen
        mask_not_in_vc = j["status"].map(lambda s: (s or "").lower() == STATUS_NOT_IN_VCENTER)
        mask_seen = mask_not_in_vc & j.apply(seen_in_openvas, axis=1)
        j.loc[mask_seen, "status"] = STATUS_ONLINE

        OLD = NOTE_PRESUME_OFFLINE
        NEW = NOTE_FOUND_VCENTER_MISS
        for idx in j.index[mask_seen]:
            notes = (j.at[idx, "notes"] or "").replace(f"; {OLD}", "").replace(OLD, "").strip()
            if NEW not in notes:
                notes = (notes + "; " + NEW).strip("; ").strip()
            j.at[idx, "notes"] = notes

        # Add net-new IPs (same as build)
        add_rows = []
        for ip in sorted(ov_ip_set):
            if ip and ip not in known_ips:
                add_rows.append({
                    "name": "",
                    "ip_address": ip,
                    "data_classification": None,
                    "type_user": None,
                    "purpose": None,
                    "dns_name": "",
                    "location": None,
                    "status": STATUS_FOUND_IN_SCAN,
                    "svc_tag": None,
                    "source": SOURCE_OPENVAS,
                    "notes": None
                })
        if add_rows:
            add = pd.DataFrame(add_rows)
            j = pd.concat([j, add], ignore_index=True, sort=False)

        # OS guess into notes only for 'found in scan'
        def enrich_notes_with_os(row):
            status = (row.get("status") or "").strip().lower()
            if status != STATUS_FOUND_IN_SCAN:
                return row.get("notes")
            existing = row.get("notes") or ""
            ip = str(row.get("ip_address") or "").lower()
            cpes = ov_cpe_map.get(ip, [])
            if cpes:
                os_guess = next((c for c in cpes if c.startswith("cpe:/o:")), None) or \
                           next((c for c in cpes if c.startswith("cpe:/a:")), None)
                if os_guess:
                    return (existing + "; " if existing else "") + f"Found OS: {os_guess}"
            return existing
        j["notes"] = j.apply(enrich_notes_with_os, axis=1)

    # --- Final sort + canonicalize
    def _ip_sort_key(val):
        try:
            parts = [int(p) for p in str(val).split(".")]
            while len(parts) < 4: parts.append(0)
            return tuple(parts[:4])
        except Exception:
            return (999, 999, 999, 999)

    j = j.copy()
    if "name" not in j.columns: j["name"] = ""
    j["_ip_sort"] = j["ip_address"].map(_ip_sort_key)
    j = j.sort_values(by=["_ip_sort", "name"], kind="mergesort").drop(columns=["_ip_sort"])

    j["status"] = j["status"].astype(str).str.strip()
    j["status"] = j["status"].replace({
        r"(?i)^powered\s*off$": "PoweredOff",
        r"(?i)^powered\s*on$": "PoweredOn",
        r"(?i)^not\s*in\s*v\s*center$": STATUS_NOT_IN_VCENTER,
        r"(?i)^not\s*in\s*vcenter$": STATUS_NOT_IN_VCENTER,
        r"(?i)^found\s*in\s*scan$": STATUS_FOUND_IN_SCAN,
        r"(?i)^online$": STATUS_ONLINE,
        r"(?i)^physical$": STATUS_PHYSICAL
    }, regex=True)

    # Build helper sheets if vCenter refreshed
    asset_ips = None
    if vc is not None:
        ip_rows = []
        for _, r in vc.iterrows():
            for ip in coerce_ip_list(r.get("ip_addresses")):
                if not ip_valid(ip): continue
                ip_rows.append({
                    "name": r.get("vc_name"),
                    "dns_name": r.get("vc_host"),
                    "ip_address": ip,
                    "vc_powerstate": r.get("vc_powerstate"),
                    "vc_vmhost": r.get("vc_vmhost"),
                    "vcenter_inventory_timestamp": r.get("vcenter_inventory_timestamp")
                })
        asset_ips = pd.DataFrame(ip_rows) if ip_rows else pd.DataFrame(
            columns=["name","dns_name","ip_address","vc_powerstate","vc_vmhost","vcenter_inventory_timestamp"]
        )

    # Meta sheet
    meta = pd.DataFrame([{ 
        "openvas_csv_path": (ov_file_path or ""),
        "vcenter_inventory_timestamp": ("" if vc_ts is None else str(vc_ts)),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "baseline": str(baseline_xlsx)
    }])

    # --- Write workbook
    with pd.ExcelWriter(output, engine="openpyxl") as w:
        # asset_tracker first
        j.to_excel(w, sheet_name="asset_tracker", index=False)

        # other baseline sheets (except vCenter_View if we are refreshing it)
        for sheet_name, df in sheets.items():
            if sheet_name == "asset_tracker": continue
            if sheet_name == "vCenter_View" and vc_tab is not None: continue
            df.to_excel(w, sheet_name=sheet_name, index=False)

        # replace vCenter sheets if refreshed
        if vc_tab is not None:
            vc_tab.to_excel(w, sheet_name="vCenter_View", index=False)
        if asset_ips is not None:
            asset_ips.to_excel(w, sheet_name="asset_ips", index=False)

        meta.to_excel(w, sheet_name="_meta", index=False)

        # Recreate filter on asset_tracker: exclude PoweredOff + not in vcenter, and hide those rows
        ws = w.book["asset_tracker"]
        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(ws.max_column)
        last_row = ws.max_row
        ws.auto_filter.ref = None
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # auto-fit
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row,
                                                         min_col=1, max_col=ws.max_column), start=1):
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len: max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

        # filter & hide
        unique_statuses = [str(s) for s in j["status"].dropna().unique().tolist()]
        blocked = {"poweredoff", STATUS_NOT_IN_VCENTER}
        allowed = [s for s in unique_statuses if s.strip().lower() not in blocked]
        status_col_idx0 = j.columns.get_loc("status")
        if allowed:
                ws.auto_filter.add_filter_column(status_col_idx0, allowed, blank=True)
                status_col_1b = status_col_idx0 + 1
                for r in range(2, last_row + 1):
                    val = ws.cell(row=r, column=status_col_1b).value
                    if isinstance(val, str) and val.strip().lower() in blocked:
                        ws.row_dimensions[r].hidden = True

    print(f"[update-asset-tracker] Wrote: {output}")

# ----------------------------
# CLI
# ----------------------------

def main():
    ap = argparse.ArgumentParser(
        description=(
            "Infrastructure Asset Tracker\n\n"
            "Commands:\n"
            "  build-asset-tracker  Build a new tracker from physical+virtual+vCenter+OpenVAS inputs.\n"
            "  update-asset-tracker Update an existing tracker using a prior XLSX and new vCenter and/or OpenVAS.\n"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    sub = ap.add_subparsers(dest="cmd")

    # BUILD (config-first, CLI can override)
    p = sub.add_parser(
        "build-asset-tracker",
        help="Build a new tracker from CSV inputs (physical, virtual, vCenter, OpenVAS).",
    )
    p.add_argument("--config", help="Path to YAML with keys: physical, virtual, vcenter, openvas_root, output")
    p.add_argument("--physical", help="Physical inventory CSV (e.g., PI24.csv)")
    p.add_argument("--virtual", help="Virtual inventory CSV (e.g., cms_cvi25_listing.csv)")
    p.add_argument("--vcenter", help="vCenter VM list CSV (e.g., vsphere_vmlist_YYYYMMDD-HHMMSS.csv)")
    p.add_argument("--openvas-root", help="Folder containing latest detailedresults.csv (OpenVAS)")
    p.add_argument("--output", help="Output XLSX path (e.g., ./outputs/Asset_Tracker.xlsx)")

    # UPDATE (no YAML required; sensible defaults for baseline/output)
    u = sub.add_parser(
        "update-asset-tracker",
        help=(
            "Update an existing tracker incrementally.\n"
            "Defaults: --baseline ./outputs/Asset_Tracker.xlsx; if --output omitted, overwrite baseline in-place (with timestamped backup)."
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    u.add_argument("--baseline", help="Existing tracker XLSX to update (default: ./outputs/Asset_Tracker.xlsx)")
    u.add_argument("--vcenter", help="New vCenter VM list CSV to apply (optional)")
    u.add_argument("--openvas-root", help="Folder containing latest detailedresults.csv (optional)")
    u.add_argument("--output", help="Output XLSX; if omitted, overwrites baseline after making a timestamped backup")

    args = ap.parse_args()

    if args.cmd == "build-asset-tracker":
        # Load config if provided
        cfg = {}
        if args.config:
            if yaml is None:
                raise RuntimeError("pyyaml is required for --config. Install with: pip install pyyaml")
            with open(args.config, "r") as f:
                data = yaml.safe_load(f) or {}
                if not isinstance(data, dict):
                    raise ValueError(f"Config file must be a YAML mapping/object, got: {type(data)}")
                cfg = data

        # Merge precedence: CLI > config
        physical     = args.physical     or cfg.get("physical")
        virtual      = args.virtual      or cfg.get("virtual")
        vcenter      = args.vcenter      or cfg.get("vcenter")
        openvas_root = args.openvas_root or cfg.get("openvas_root")
        output       = args.output       or cfg.get("output")

        # Validate required fields after merge
        missing = [k for k, v in {
            "physical": physical,
            "virtual": virtual,
            "vcenter": vcenter,
            "openvas_root": openvas_root,
            "output": output
        }.items() if not v]
        if missing:
            ap.error(f"Missing required option(s) after merging with config: {', '.join(missing)}")

        build_asset_tracker(
            Path(physical),
            Path(virtual),
            Path(vcenter),
            Path(openvas_root),
            Path(output),
        )
        return

    if args.cmd == "update-asset-tracker":
        baseline = Path(args.baseline) if args.baseline else Path("./outputs/Asset_Tracker.xlsx")
        vcenter  = Path(args.vcenter) if args.vcenter else None
        openvas  = Path(args.openvas_root) if args.openvas_root else None
        output   = Path(args.output) if args.output else None  # None => overwrite baseline (with backup)

        if not vcenter and not openvas:
            ap.error("Provide at least one of --vcenter or --openvas-root for update.")

        update_asset_tracker(baseline, vcenter, openvas, output)
        return

    ap.print_help()

if __name__ == "__main__":
    main()
